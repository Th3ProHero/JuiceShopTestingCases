import unittest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from common import RegistrationPage
from config import DRIVER_PATH

class RegistrationTest(unittest.TestCase):

    def setUp(self):
        self.service = Service(DRIVER_PATH)

    def test_registration_success(self):
        wb = load_workbook('users.xlsx')
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 3:
                print(f"Fila {row_idx} incompleta, saltando...")
                continue

            email, password, registrado = row
            if registrado != "Sí":
                self.driver = webdriver.Chrome(service=self.service)
                self.driver.get("http://192.168.1.68:4000/#/register")
                self.registration_page = RegistrationPage(self.driver)
                
                try:
                    selected_question = self.registration_page.register(
                        email=email,
                        password=password,
                        confirm_password=password,
                        answer="John"
                    )
                    # Esperar hasta que el mensaje de bienvenida esté presente
                    WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/div/mat-sidenav-container/mat-sidenav-content/app-login/div/mat-card/h1"))
                    )
                    ws.cell(row=row_idx, column=3, value="Sí")
                except Exception as e:
                    ws.cell(row=row_idx, column=3, value="No")
                    print(f"Error registrando {email}: {e}")
                finally:
                    self.driver.quit()

        wb.save('users.xlsx')

    def tearDown(self):
        pass  # El navegador se cierra en cada iteración del bucle

if __name__ == "__main__":
    unittest.main()
